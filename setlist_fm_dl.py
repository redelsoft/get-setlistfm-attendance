#!/usr/bin/env python3

import requests
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Set your API key and the base URL
api_key = '<API_KEY>'
base_url = 'https://api.setlist.fm/rest/1.0/'

# Set your username and headers
username = '<USERNAME>'
headers = {
    'Accept': 'application/json',
    'x-api-key': api_key
}


def get_user_setlists(username):
    setlists = []
    page = 1

    while True:
        url = f'{base_url}user/{username}/attended?p={page}'
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            setlists.extend(data.get('setlist', []))  # Add current page results to the list
            
            total = data.get('total', 0)
            items_per_page = data.get('itemsPerPage', 0)
            
            if page * items_per_page >= total:
                break  # All pages retrieved
            else:
                page += 1  # Move to the next page
        else:
            print(f"Error: {response.status_code}, {response.text}")
            break
        
        sleep(2)

    return setlists


def write_to_excel(setlists, username, filename=None):
    """
    Create an Excel workbook from a list of concert setlists and save it to a specified filename.

    Args:
        setlists (list): A list of dictionaries containing concert details.
        username (str): The username for the Excel sheet title.
        filename (str, optional): The name of the output Excel file. Defaults to f'concerts_{username}.xlsx'.
    """
    if filename is None:
        filename = f'concerts_{username}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = f"Concerts {username}"

    headers = ['Artist', 'Date', 'Venue', 'City', 'Country']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")

    for idx, setlist in enumerate(setlists, start=2):
        artist = setlist['artist']['name']
        date = setlist['eventDate']
        venue = setlist['venue']['name']
        city = setlist['venue']['city']['name']
        country = setlist['venue']['city']['country']['name']

        row = [artist, date, venue, city, country]
        ws.append(row)

        if idx % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws[f'{get_column_letter(col_num)}{idx}'].fill = PatternFill(start_color="FFF0F0F0",
                                                                            end_color="FFF0F0F0", fill_type="solid")

    ws.auto_filter.ref = ws.dimensions

    for col_num in range(1, len(headers) + 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(col_num)])
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    wb.save(filename)
    print(f"Data written to {filename}")


def main():
    setlists = get_user_setlists(username)
    
    if setlists:
        write_to_excel(setlists, username=username)
    else:
        print("No concerts found.")


if __name__ == "__main__":
    main()
